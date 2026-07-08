#!/usr/bin/env python3
"""
Political Narrative Discovery Pipeline  v4.5 (FINAL with Semantic Normalization)
=====================================================
Unsupervised discovery of HUMAN-QUALITY political narratives from tweets.

Key Upgrades:
  • Intelligent noise re-assignment via tier-based approach
  • 2-Stage Clustering: Merges fragmented HDBSCAN clusters using summary embeddings
  • Upgraded to Flan-T5 Large for production-quality summaries
  • Direct LLM prompting for human-like narrative title generation
  • Optimized Excel export (cleaner columns, specific widths, polished styling)
  • Logarithmic Impact Score calculation and cluster-level aggregations
  • SpaCy NER + KeyBERT for intelligent, deduplicated Key Themes extraction
  • Pairwise cluster coherence scoring and console Quality Report
  • Semantic Normalization layer to strip repetitive framing noise from embeddings

Pipeline:
  Tweets → Cleaning → Semantic Normalization → BGE Embedding → UMAP → HDBSCAN 
    → Noise Re-assignment → Base Summaries → Embed Summaries → Merge Similar Narratives 
    → Final Abstractive Summary → LLM Title Generation → Key Themes 
    → Impact Scoring → Coherence Check → Quality Report

Output: 3 Excel sheets
  1. tweets_clustered: Full data + cluster + narrative + impact_score
  2. narrative_summary: Cluster aggregates (counts, %, and engagement metrics)
  3. cluster_summary: Full diagnostics with Key Themes and cluster summaries

Usage:
    python narrative_discovery_final.py <input_excel_path> [output_excel_path]

Dependencies:
    pip install pandas openpyxl numpy scikit-learn umap-learn hdbscan
    pip install sentence-transformers transformers torch
    pip install spacy keybert
    python -m spacy download en_core_web_sm
"""

import sys
import re
import warnings
import unicodedata
from pathlib import Path
from typing import Tuple, Dict, List, Set

import numpy as np
import pandas as pd
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

warnings.filterwarnings("ignore")

# =============================================================================
# CONFIGURATION
# =============================================================================

# --- Embedding model ---
EMBEDDING_MODEL = "BAAI/bge-base-en-v1.5"

# --- Local summarization model ---
SUMMARIZATION_MODEL = "google/flan-t5-large"

# --- UMAP Configuration ---
UMAP_CONFIG = {
    "metric": "cosine",
    "min_dist": 0.0,
    "random_state": 42,
}

# --- HDBSCAN Configuration ---
HDBSCAN_CONFIG = {
    "metric": "euclidean",
    "cluster_selection_method": "leaf",
}

# --- Clustering / Merging Thresholds ---
NOISE_SIMILARITY_THRESHOLD = 0.70  # Min cosine similarity to assign noise to an existing cluster
MERGE_SIMILARITY_THRESHOLD = 0.85  # Min cosine similarity between summaries to merge clusters

# --- Narrative Generation ---
TFIDF_TOP_N_TERMS    = 10
CENTRALITY_TOP_N     = 5

# --- Output ---
OUTPUT_FONT = "Arial"


# =============================================================================
# GLOBAL CACHES
# =============================================================================

_SUMMARIZER_CACHE = None
_EMBEDDING_MODEL_CACHE = None
_SPACY_MODEL_CACHE = None
_KEYBERT_MODEL_CACHE = None

def _load_embedding_model(model_name: str):
    global _EMBEDDING_MODEL_CACHE
    if _EMBEDDING_MODEL_CACHE is None:
        from sentence_transformers import SentenceTransformer
        print(f"  Loading embedding model: {model_name}")
        _EMBEDDING_MODEL_CACHE = SentenceTransformer(model_name)
    return _EMBEDDING_MODEL_CACHE


def _load_summarizer(model_name: str):
    global _SUMMARIZER_CACHE
    if _SUMMARIZER_CACHE is not None:
        return _SUMMARIZER_CACHE

    from transformers import pipeline
    print(f"\n  Loading summarization model: {model_name}")
    
    if "flan-t5" in model_name.lower():
        _SUMMARIZER_CACHE = pipeline(
            "text2text-generation",
            model=model_name,
            max_new_tokens=80,
            do_sample=False,
        )
    else:
        _SUMMARIZER_CACHE = pipeline(
            "summarization",
            model=model_name,
            max_length=80,
            min_length=20,
            do_sample=False,
        )
    return _SUMMARIZER_CACHE


def _load_spacy():
    global _SPACY_MODEL_CACHE
    if _SPACY_MODEL_CACHE is None:
        import spacy
        try:
            _SPACY_MODEL_CACHE = spacy.load("en_core_web_sm")
        except OSError:
            print("\n  [Setup] Downloading spaCy model 'en_core_web_sm'...")
            import subprocess
            subprocess.run([sys.executable, "-m", "spacy", "download", "en_core_web_sm"], check=True)
            _SPACY_MODEL_CACHE = spacy.load("en_core_web_sm")
    return _SPACY_MODEL_CACHE


def _load_keybert():
    global _KEYBERT_MODEL_CACHE
    if _KEYBERT_MODEL_CACHE is None:
        from keybert import KeyBERT
        _KEYBERT_MODEL_CACHE = KeyBERT(model=_load_embedding_model(EMBEDDING_MODEL))
    return _KEYBERT_MODEL_CACHE


# =============================================================================
# PART 1: ADAPTIVE PARAMETER HELPERS
# =============================================================================

def get_umap_config(n_tweets: int) -> dict:
    cfg = UMAP_CONFIG.copy()
    if n_tweets < 30:
        cfg["n_neighbors"] = max(2, n_tweets - 1)
        cfg["n_components"] = min(3, n_tweets - 2)
    elif n_tweets < 100:
        cfg["n_neighbors"] = 10
        cfg["n_components"] = 5
    elif n_tweets < 500:
        cfg["n_neighbors"] = 15
        cfg["n_components"] = 8
    else:
        cfg["n_neighbors"] = 30
        cfg["n_components"] = 10
    return cfg


def get_hdbscan_config(n_tweets: int) -> dict:
    cfg = HDBSCAN_CONFIG.copy()
    if n_tweets < 30:
        cfg["min_cluster_size"] = 2
        cfg["min_samples"] = 1
    elif n_tweets < 100:
        cfg["min_cluster_size"] = 3
        cfg["min_samples"] = 2
    elif n_tweets < 500:
        cfg["min_cluster_size"] = 4
        cfg["min_samples"] = 2
    else:
        cfg["min_cluster_size"] = 8
        cfg["min_samples"] = 4
    return cfg


# =============================================================================
# PART 2: DATA LOADING, CLEANING, NORMALIZATION & IMPACT SCORING
# =============================================================================

def load_data(excel_path: str) -> pd.DataFrame:
    path = Path(excel_path)
    if path.is_dir():
        files = sorted(path.glob("*.xlsx"))
        if not files:
            raise FileNotFoundError(f"No .xlsx files found in: {excel_path}")
        frames = [pd.read_excel(f) for f in files]
        df = pd.concat(frames, ignore_index=True)
        print(f"  Merged {len(files)} files → {len(df)} total rows")
    else:
        df = pd.read_excel(excel_path)
        print(f"  Loaded: {excel_path} → {len(df)} rows")

    if "tweet_text" not in df.columns:
        raise ValueError(f"'tweet_text' column required. Got: {df.columns.tolist()}")

    df = df.dropna(subset=["tweet_text"]).reset_index(drop=True)
    df["tweet_text"] = df["tweet_text"].astype(str)
    return df


_BOLD_MAP = {
    **{chr(0x1D400 + i): chr(ord('A') + i) for i in range(26)},
    **{chr(0x1D41A + i): chr(ord('a') + i) for i in range(26)},
    **{chr(0x1D468 + i): chr(ord('A') + i) for i in range(26)},
    **{chr(0x1D482 + i): chr(ord('a') + i) for i in range(26)},
    **{chr(0x1D7CE + i): str(i) for i in range(10)},
}


def clean_text(text: str) -> str:
    if not isinstance(text, str): return ""
    result = "".join([_BOLD_MAP.get(ch, unicodedata.normalize("NFKC", ch)) for ch in text])
    text = re.sub(r"https?://\S+|www\.\S+", " ", result)
    text = re.sub(r"#(\w+)", r"\1", text)
    text = re.sub(r"@(\w+)", r"\1", text)
    text = re.sub(r"[^\w\s\-'',\.:%₹\(\)/]", " ", text)
    text = re.sub(r"[\n\r\t]+", " ", text)
    return re.sub(r" {2,}", " ", text).strip()


def clean_corpus(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["cleaned_text"] = df["tweet_text"].apply(clean_text)
    empty = df["cleaned_text"].str.strip() == ""
    return df[~empty].reset_index(drop=True)


def semantic_normalize_text(text: str) -> str:
    """Extract a concise semantic representation of the tweet by stripping away repetitive

    political campaign framing and promotional prefixes while fully preserving core entities.
    """
    if not text:
        return ""
    
    # Structural case-insensitive framing patterns to look for and remove
    patterns = [
        r"\bunder the leadership of (?:pm\s+)?(?:narendra\s+)?modi(?:ji)?\b",
        r"\b(?:pm\s+)?(?:narendra\s+)?modi(?:ji)?\s+said\b",
        r"\bover the last \d+(?:/\d+)?\s+years\b",
        r"\bin the last \d+(?:/\d+)?\s+years\b",
        r"\bnew india\b",
        r"\bunder this government\b",
        r"\bgovernment led by\b",
        r"\bvision of (?:pm\s+)?(?:narendra\s+)?modi\b",
        r"\bled by (?:pm\s+)?(?:narendra\s+)?modi\b",
        r"\bunder the visionary leadership of\b",
        r"\bunder the guidance of\b",
        r"\btransformational journey of\b",
    ]
    
    normalized = text
    for pattern in patterns:
        normalized = re.sub(pattern, "", normalized, flags=re.IGNORECASE)
        
    # Clean up multiple whitespaces
    normalized = re.sub(r"\s+", " ", normalized).strip()
    
    # Strip common dangling syntax leftovers at the very start of a sentence after deletions
    normalized = re.sub(r"^(?:that|has|have|had|is|are|was|were|,|\s)+", "", normalized, flags=re.IGNORECASE)
    
    # Clean syntax punctuation bounds
    normalized = re.sub(r"^[,\.:;\-\u2014\s]+", "", normalized)
    normalized = re.sub(r"[,\.:;\-\u2014\s]+$", "", normalized)
    normalized = normalized.strip()
    
    # Fallback safety validation
    if not normalized:
        return text.strip()
    return normalized


def calculate_impact_scores(df: pd.DataFrame) -> pd.DataFrame:
    """Calculate normalized logarithmic impact score based on engagement metrics."""
    df = df.copy()
    
    metrics = ["view_count", "like_count", "repost_count", "reply_count", "bookmark_count"]
    for m in metrics:
        if m not in df.columns:
            df[m] = 0
        else:
            df[m] = pd.to_numeric(df[m], errors="coerce").fillna(0)

    views_score = np.minimum(100, 100 * np.log10(df["view_count"] + 1) / np.log10(500000))
    likes_score = np.minimum(100, 100 * np.log10(df["like_count"] + 1) / np.log10(10000))
    repost_score = np.minimum(100, 100 * np.log10(df["repost_count"] + 1) / np.log10(3000))
    reply_score = np.minimum(100, 100 * np.log10(df["reply_count"] + 1) / np.log10(300))
    bookmark_score = np.minimum(100, 100 * np.log10(df["bookmark_count"] + 1) / np.log10(150))

    df["impact_score"] = (
        0.30 * views_score +
        0.30 * likes_score +
        0.20 * repost_score +
        0.10 * reply_score +
        0.10 * bookmark_score
    ).round(2)

    return df


# =============================================================================
# PART 3: EMBEDDINGS (BGE + Cosine Normalization)
# =============================================================================

def generate_embeddings(texts: List[str], model_name: str = EMBEDDING_MODEL) -> np.ndarray:
    model = _load_embedding_model(model_name)
    print(f"  Encoding {len(texts)} chunks...")
    embeddings = model.encode(
        texts,
        batch_size=32,
        show_progress_bar=False,
        convert_to_numpy=True,
        normalize_embeddings=True,
    )
    return embeddings


# =============================================================================
# PART 4 & 5: REDUCTION & CLUSTERING (UMAP + HDBSCAN + Noise Handler)
# =============================================================================

def reduce_dimensions(embeddings: np.ndarray, n_tweets: int) -> np.ndarray:
    import umap as umap_module
    cfg = get_umap_config(n_tweets)
    reducer = umap_module.UMAP(**cfg)
    return reducer.fit_transform(embeddings)


def reassign_noise_points(labels: np.ndarray, embeddings: np.ndarray) -> np.ndarray:
    """Assign noise (-1) to the nearest cluster using a tier-based approach."""
    valid_idx = labels >= 0
    if not valid_idx.any():
        return labels

    unique_clusters = set(labels[valid_idx])
    c_ids = list(unique_clusters)
    centroids = []
    
    for c in c_ids:
        mask = labels == c
        centroid = embeddings[mask].mean(axis=0)
        centroids.append(centroid / (np.linalg.norm(centroid) + 1e-9))
    
    centroid_mat = np.vstack(centroids)
    new_labels = labels.copy()
    next_id = int(labels.max()) + 1
    
    noise_indices = np.where(labels == -1)[0]
    reassigned = 0
    
    cluster_counts = {c: int((labels == c).sum()) for c in c_ids}
    
    for i in noise_indices:
        sims = cosine_similarity(embeddings[i:i+1], centroid_mat)[0]
        best_idx = np.argmax(sims)
        best_sim = sims[best_idx]
        best_cluster = c_ids[best_idx]
        
        if best_sim >= 0.78:
            new_labels[i] = best_cluster
            reassigned += 1
        elif 0.68 <= best_sim < 0.78 and cluster_counts[best_cluster] >= 5:
            new_labels[i] = best_cluster
            reassigned += 1
        else:
            new_labels[i] = next_id
            next_id += 1
            
    print(f"  Reassigned {reassigned} noise points to existing clusters via Tiered rules.")
    print(f"  Created {len(noise_indices) - reassigned} new singletons.")
    return new_labels


def cluster_embeddings(reduced: np.ndarray, embeddings: np.ndarray, n_tweets: int) -> np.ndarray:
    import hdbscan
    cfg = get_hdbscan_config(n_tweets)
    
    clusterer = hdbscan.HDBSCAN(**cfg)
    labels = clusterer.fit_predict(reduced)
    
    n_real = len(set(labels)) - (1 if -1 in labels else 0)
    n_noise = int((labels == -1).sum())
    print(f"  HDBSCAN raw clusters: {n_real} | noise points: {n_noise}")
    
    if n_noise > 0:
        labels = reassign_noise_points(labels, embeddings)
        
    return labels


# =============================================================================
# PART 6: NARRATIVE GENERATION (2-Stage Merging + LLM Titles + Key Themes)
# =============================================================================

def _stopwords() -> Set[str]:
    return {
        "the","a","an","and","or","but","in","on","at","to","for","of","with",
        "by","from","is","are","was","were","be","been","has","have","had",
        "will","would","can","could","should","do","does","did","not","no",
        "it","its","this","that","these","those","they","their","them","we",
        "our","you","your","he","she","his","her","as","so","if","how","what",
        "when","where","who","which","while","after","before","more","also",
        "just","now","new","into","out","up","down","over","under","than","then",
        "there","here","about","like","through","across","well","all","one",
        "two","three","many","much","every","each","rt","via","amp","etc",
        "per","https","http","www","today","india","indian","pm","shri",
        "narendramodi","bjp4india","modi",
    }


def _tfidf_keywords(texts: List[str], top_n: int = TFIDF_TOP_N_TERMS) -> List[str]:
    stop = _stopwords()
    if len(texts) < 2:
        words = [w for w in re.findall(r'\b[A-Za-z][A-Za-z0-9\-]{2,}\b', " ".join(texts)) if w.lower() not in stop]
        freq = {w: words.count(w) for w in set(words)}
        return sorted(freq, key=freq.get, reverse=True)[:top_n]
    try:
        vec = TfidfVectorizer(max_features=400, stop_words=list(stop), ngram_range=(1, 2), token_pattern=r'\b[A-Za-z][A-Za-z0-9\-]{2,}\b')
        mat = vec.fit_transform(texts)
        scores = np.asarray(mat.mean(axis=0)).flatten()
        top = scores.argsort()[::-1][:top_n]
        return [vec.get_feature_names_out()[i] for i in top]
    except Exception:
        return []


def _extract_key_themes(texts: List[str], summary: str, top_n: int = 8) -> List[str]:
    nlp = _load_spacy()
    kw_model = _load_keybert()
    
    doc_text = summary + " " + " ".join(texts)
    if len(doc_text) > 500000:
        doc_text = doc_text[:500000]
        
    doc = nlp(doc_text)
    valid_ents = {"ORG", "PERSON", "GPE", "LOC", "EVENT", "PRODUCT", "WORK_OF_ART", "LAW", "NORP"}
    entity_freq = {}
    
    for ent in doc.ents:
        if ent.label_ in valid_ents:
            t = re.sub(r"[^a-zA-Z0-9\s]", "", ent.text).strip().title()
            if len(t) > 2:
                entity_freq[t] = entity_freq.get(t, 0) + 1
                
    top_entities = [e for e, _ in sorted(entity_freq.items(), key=lambda x: x[1], reverse=True)[:10]]
    
    central_text = summary + "\n" + "\n".join(texts[:100])
    keywords = kw_model.extract_keywords(
        central_text,
        keyphrase_ngram_range=(1, 3),
        stop_words="english",
        use_mmr=True,
        diversity=0.4,
        top_n=15
    )
    kb_phrases = [kw.strip().title() for kw, score in keywords if len(kw.strip()) > 2]
    
    generic_words = _stopwords() | {
        "story", "year", "once", "people", "must", "time", "day", "week", "month", "today",
        "tomorrow", "yesterday", "many", "much", "video", "photo", "image", "tweet", "tweets",
        "update", "news", "latest", "watch", "look", "good", "great", "best", "like", "share",
        "things", "something", "anything", "nothing", "everything", "someone", "anyone", "everyone",
        "know", "think", "want", "need", "make", "take", "come", "give", "work"
    }
    
    merged_themes = []
    raw_pool = top_entities + kb_phrases
    
    for item in raw_pool:
        item_lower = item.lower()
        if item.isnumeric() or item_lower in generic_words:
            continue
            
        words = item_lower.split()
        if all(w in generic_words for w in words):
            continue

        is_subset = False
        to_remove = None
        
        for existing in merged_themes:
            ex_lower = existing.lower()
            if item_lower == ex_lower or item_lower in ex_lower:
                is_subset = True
                break
            if ex_lower in item_lower:
                to_remove = existing
                break
        
        if to_remove:
            merged_themes.remove(to_remove)
            merged_themes.append(item)
        elif not is_subset:
            merged_themes.append(item)
            
    return merged_themes[:top_n]


def _central_tweets(texts: List[str], embeddings_subset: np.ndarray, top_n: int = CENTRALITY_TOP_N) -> List[str]:
    if len(texts) == 1: return texts
    centroid = embeddings_subset.mean(axis=0, keepdims=True)
    centroid /= (np.linalg.norm(centroid) + 1e-9)
    sims = cosine_similarity(centroid, embeddings_subset)[0]
    top_idx = sims.argsort()[::-1][:min(top_n, len(texts))]
    return [texts[i] for i in top_idx]


def _generate_cluster_summary(central_tweets: List[str], keywords: List[str], summarizer_model: str) -> str:
    summarizer = _load_summarizer(summarizer_model)
    tweet_block = "\n".join(f"- {t[:300]}" for t in central_tweets)
    kw_hint = ", ".join(keywords[:6])

    if "flan-t5" in summarizer_model.lower():
        prompt = (
            f"Summarize the main political theme or event discussed in these tweets in one clear sentence. "
            f"Key topics include: {kw_hint}.\n\n"
            f"Tweets:\n{tweet_block}\n\n"
            f"One-sentence summary:"
        )
        summary = summarizer(prompt)[0]["generated_text"].strip()
    else:
        concat = " ".join(t[:200] for t in central_tweets)[:1024]
        summary = summarizer(concat)[0]["summary_text"].strip()

    return summary if len(summary) > 10 else f"Political content related to: {kw_hint}"


def _generate_narrative_title(summary: str, summarizer_model: str) -> str:
    summarizer = _load_summarizer(summarizer_model)
    
    if "flan-t5" in summarizer_model.lower():
        prompt = (
            f"Create a short, punchy 3 to 5 word political narrative title for the following summary.\n"
            f"Summary: {summary}\n\n"
            f"Narrative Title:"
        )
        title = summarizer(prompt, max_new_tokens=15)[0]["generated_text"].strip()
    else:
        title = summarizer(summary, max_length=12, min_length=2)[0]["summary_text"].strip()
        
    title = title.strip('"\'.')
    return " ".join([w.capitalize() for w in title.split()[:7]])


def assign_narrative_labels(df: pd.DataFrame, embeddings: np.ndarray, summarizer_model: str) -> Tuple[pd.DataFrame, Dict]:
    df = df.copy()
    
    # ---------------------------------------------------------
    # STAGE 1: Generate base summaries for fragment clusters
    # ---------------------------------------------------------
    base_clusters = sorted(df["cluster"].unique())
    base_summaries = {}
    print("\n  [6a] Generating base summaries for narrative merging...")
    
    for cid in base_clusters:
        mask = df["cluster"] == cid
        texts = df.loc[mask, "cleaned_text"].tolist()
        embs = embeddings[mask.values]
        rep_tweets = _central_tweets(texts, embs, top_n=3)
        base_summaries[cid] = _generate_cluster_summary(rep_tweets, [], summarizer_model)

    # ---------------------------------------------------------
    # STAGE 2: Embed summaries and merge highly similar ones
    # ---------------------------------------------------------
    print("  [6b] Embedding summaries & merging similar narratives...")
    sum_texts = [base_summaries[cid] for cid in base_clusters]
    sum_embs = generate_embeddings(sum_texts, EMBEDDING_MODEL)
    
    sim_mat = cosine_similarity(sum_embs)
    np.fill_diagonal(sim_mat, 0)
    
    parent = {c: c for c in base_clusters}
    def find(i):
        if parent[i] == i: return i
        parent[i] = find(parent[i])
        return parent[i]
        
    for i in range(len(base_clusters)):
        for j in range(i+1, len(base_clusters)):
            if sim_mat[i, j] > MERGE_SIMILARITY_THRESHOLD:
                root_i, root_j = find(base_clusters[i]), find(base_clusters[j])
                if root_i != root_j:
                    parent[root_i] = root_j
                    
    mapping = {c: find(c) for c in base_clusters}
    df["cluster"] = df["cluster"].map(mapping)
    
    # ---------------------------------------------------------
    # STAGE 3: Finalize Generation on Merged Clusters
    # ---------------------------------------------------------
    print("  [6c] Generating final overarching narratives, titles & Key Themes...")
    final_clusters = sorted(df["cluster"].unique())
    cluster_to_info = {}
    
    for idx, cid in enumerate(final_clusters):
        mask = df["cluster"] == cid
        texts = df.loc[mask, "cleaned_text"].tolist()
        embs = embeddings[mask.values]

        if len(embs) > 1:
            sim_matrix = cosine_similarity(embs)
            np.fill_diagonal(sim_matrix, -1.0)
            max_sims = sim_matrix.max(axis=1)
            coherence_score = float(max_sims.mean())
        else:
            coherence_score = 0.0

        hint_keywords = _tfidf_keywords(texts, top_n=TFIDF_TOP_N_TERMS)
        rep_tweets = _central_tweets(texts, embs, top_n=CENTRALITY_TOP_N)
        summary = _generate_cluster_summary(rep_tweets, hint_keywords, summarizer_model)
        title = _generate_narrative_title(summary, summarizer_model)

        final_key_themes = _extract_key_themes(texts, summary, top_n=8)

        cluster_to_info[cid] = {
            "narrative": title,
            "summary": summary,
            "keywords": final_key_themes,
            "rep_tweets": rep_tweets,
            "coherence_score": coherence_score,
        }
        print(f"    [{idx+1}/{len(final_clusters)}] Cluster {cid:>3} ({len(texts):>3} tweets) -> {title}")

    df["narrative"] = df["cluster"].map(lambda c: cluster_to_info[c]["narrative"])
    df["cluster_summary"] = df["cluster"].map(lambda c: cluster_to_info[c]["summary"])
    return df, cluster_to_info


# =============================================================================
# PART 7: EXCEL EXPORT
# =============================================================================

def export_to_excel(df: pd.DataFrame, output_path: str, cluster_to_info: Dict, input_stem: str = "") -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    hdr_fill   = PatternFill("solid", start_color="1F3864", end_color="1F3864")
    hdr_font   = Font(name=OUTPUT_FONT, bold=True, color="FFFFFF", size=11)
    hdr_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin       = Side(style="thin", color="CCCCCC")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_alt   = PatternFill("solid", start_color="F5F7FA", end_color="F5F7FA")
    fill_white = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")

    def write_header(ws, cols):
        for ci, name in enumerate(cols, 1):
            c = ws.cell(row=1, column=ci, value=name)
            c.font, c.fill, c.alignment, c.border = hdr_font, hdr_fill, hdr_align, border
        ws.row_dimensions[1].height = 30

    def native(val):
        if isinstance(val, (np.integer,)): return int(val)
        if isinstance(val, (np.floating,)): return float(val)
        try:
            if pd.isna(val): return ""
        except Exception:
            pass
        return val

    # Sheet 1: tweets_clustered
    ws1 = wb.active
    ws1.title = "tweets_clustered"

    orig_cols = [c for c in df.columns if c not in ("cleaned_text", "cluster", "narrative", "cluster_summary", "impact_score")]
    export_cols = orig_cols + ["cluster", "narrative", "impact_score"]
    write_header(ws1, export_cols)

    narr_ci = export_cols.index("narrative") + 1

    for ri, (_, row) in enumerate(df[export_cols].iterrows(), start=2):
        fill = fill_alt if ri % 2 == 0 else fill_white
        for ci, col in enumerate(export_cols, start=1):
            val = native(row[col])
            cell = ws1.cell(row=ri, column=ci, value=val)
            cell.fill = fill
            cell.border = border
            cell.font = Font(name=OUTPUT_FONT, size=10)
            
            if col == "tweet_text":
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="center")
            
            if ci == narr_ci:
                cell.font = Font(name=OUTPUT_FONT, size=10, bold=True, color="1F3864")

    col_widths = {
        "tweet_id": 20, "datetime": 18, "tweet_url": 35, "account": 15,
        "tweet_text": 75, "reply_count": 12, "repost_count": 12,
        "like_count": 12, "bookmark_count": 14, "view_count": 12,
        "has_image": 11, "has_video": 11, "hashtags": 25, "mentions": 20,
        "language": 10, "cluster": 10, "narrative": 45, "impact_score": 15,
    }
    for ci, col in enumerate(export_cols, 1):
        ws1.column_dimensions[get_column_letter(ci)].width = col_widths.get(col, 15)
        
    ws1.freeze_panes = "A2"

    # Sheet 2: narrative_summary
    ws2 = wb.create_sheet("narrative_summary")
    total = len(df)
    
    aggs = df.groupby(["cluster", "narrative"]).agg(
        tweet_count=("cluster", "size"),
        average_impact_score=("impact_score", "mean"),
        total_views=("view_count", "sum"),
        total_likes=("like_count", "sum"),
        total_reposts=("repost_count", "sum"),
        total_replies=("reply_count", "sum"),
        total_bookmarks=("bookmark_count", "sum")
    ).reset_index()

    aggs = aggs.sort_values("tweet_count", ascending=False).reset_index(drop=True)
    aggs["percentage_of_total"] = (aggs["tweet_count"] / total * 100).round(1)
    aggs["average_impact_score"] = aggs["average_impact_score"].round(2)

    s_cols = ["cluster", "narrative", "tweet_count", "percentage_of_total",
              "average_impact_score", "total_views", "total_likes", 
              "total_reposts", "total_replies", "total_bookmarks"]
    
    write_header(ws2, s_cols)

    for ri, (_, row) in enumerate(aggs.iterrows(), start=2):
        fill = fill_alt if ri % 2 == 0 else fill_white
        for ci, col in enumerate(s_cols, 1):
            val = native(row[col])
            cell = ws2.cell(row=ri, column=ci, value=val)
            cell.font = Font(name=OUTPUT_FONT, size=10, bold=(col == "narrative"))
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if col == "percentage_of_total":
                cell.number_format = '0.0"%"'

    ws2.column_dimensions["A"].width = 10
    ws2.column_dimensions["B"].width = 45
    ws2.column_dimensions["C"].width = 15
    ws2.column_dimensions["D"].width = 22
    ws2.column_dimensions["E"].width = 22
    ws2.column_dimensions["F"].width = 15
    ws2.column_dimensions["G"].width = 15
    ws2.column_dimensions["H"].width = 15
    ws2.column_dimensions["I"].width = 15
    ws2.column_dimensions["J"].width = 18
    ws2.freeze_panes = "A2"

    # Sheet 3: cluster_summary
    ws3 = wb.create_sheet("cluster_summary")

    d_cols = ["cluster", "narrative", "cluster_summary", "top_keywords"]

    diag_hdr_fill = PatternFill("solid", start_color="14375E", end_color="14375E")
    for ci, name in enumerate(d_cols, 1):
        c = ws3.cell(row=1, column=ci, value=name)
        c.font = Font(name=OUTPUT_FONT, bold=True, color="FFFFFF", size=11)
        c.fill, c.alignment, c.border = diag_hdr_fill, hdr_align, border
    ws3.row_dimensions[1].height = 30

    accent_fills = [
        PatternFill("solid", start_color="EAF4FB", end_color="EAF4FB"),
        PatternFill("solid", start_color="FEF9E7", end_color="FEF9E7"),
        PatternFill("solid", start_color="EAFAF1", end_color="EAFAF1"),
        PatternFill("solid", start_color="FDEDEC", end_color="FDEDEC"),
        PatternFill("solid", start_color="F5EEF8", end_color="F5EEF8"),
    ]

    sorted_clusters = (
        df.groupby("cluster").size()
          .reset_index(name="n")
          .sort_values("n", ascending=False)["cluster"]
          .tolist()
    )

    for ri, cid in enumerate(sorted_clusters, start=2):
        info = cluster_to_info[cid]
        
        kws = " • ".join(info["keywords"])
        rfill = accent_fills[(ri - 2) % len(accent_fills)]

        row_vals = [cid, info["narrative"], info["summary"], kws]

        for ci, val in enumerate(row_vals, 1):
            cell = ws3.cell(row=ri, column=ci, value=val)
            cell.fill = rfill
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(name=OUTPUT_FONT, size=10, bold=(ci == 2), italic=(ci == 3))

    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 45
    ws3.column_dimensions["C"].width = 75
    ws3.column_dimensions["D"].width = 50
    ws3.freeze_panes = "A2"

    wb.save(output_path)
    print(f"\n  ✓ Output saved → {output_path}")


# =============================================================================
# PART 8: DIAGNOSTICS & QUALITY REPORT
# =============================================================================

def print_cluster_quality_report(df: pd.DataFrame, cluster_to_info: Dict) -> None:
    """Generates and prints a detailed evaluation summary of final clusters."""
    print("\n" + "=" * 95)
    print("CLUSTER QUALITY REPORT")
    print("=" * 95)
    print(f"{'Cluster ID':<12} {'Narrative Title':<42} {'Tweet Count':<14} {'Coherence':<12} {'Quality Flag'}")
    print("-" * 95)
    
    sorted_clusters = (
        df.groupby("cluster").size()
          .reset_index(name="n")
          .sort_values("n", ascending=False)["cluster"]
          .tolist()
    )
    
    for cid in sorted_clusters:
        info = cluster_to_info[cid]
        count = int((df["cluster"] == cid).sum())
        coherence = info.get("coherence_score", 0.0)
        
        if count == 1:
            flag = "•"
            coherence_str = "0.00"
        else:
            coherence_str = f"{coherence:.4f}"
            if coherence >= 0.50:
                flag = "✓ Good"
            else:
                flag = "⚠ Needs Review"
                
        print(f"{cid:<12} {info['narrative'][:40]:<42} {count:<14} {coherence_str:<12} {flag}")
    print("=" * 95 + "\n")


# =============================================================================
# MAIN PIPELINE
# =============================================================================

def run_pipeline(input_path: str, output_path: str = None) -> pd.DataFrame:
    if output_path is None:
        input_p = Path(input_path)
        output_path = str(input_p.parent / f"{input_p.stem}_narrative_output.xlsx")

    print("\n" + "=" * 70)
    print("POLITICAL NARRATIVE DISCOVERY PIPELINE  v4.5")
    print("=" * 70)

    print("\n[1/7] Loading data...")
    df = load_data(input_path)

    print("\n[2/7] Cleaning text & Calculating Impact Scores...")
    df = clean_corpus(df)
    df = calculate_impact_scores(df)
    n = len(df)

    print("\n[3/7] Generating BGE embeddings (with Semantic Normalization layer)...")
    # NEW: Isolate semantic normalization layer immediately before vector encoding
    normalized_texts = [semantic_normalize_text(t) for t in df["cleaned_text"].tolist()]
    embeddings = generate_embeddings(normalized_texts, EMBEDDING_MODEL)

    print("\n[4/7] UMAP dimensionality reduction...")
    reduced = reduce_dimensions(embeddings, n)

    print("\n[5/7] HDBSCAN clustering & Noise Re-assignment...")
    labels = cluster_embeddings(reduced, embeddings, n)
    df["cluster"] = labels

    print("\n[6/7] Generating and Merging narratives (LLM pass)...")
    df, cluster_to_info = assign_narrative_labels(df, embeddings, SUMMARIZATION_MODEL)

    print_cluster_quality_report(df, cluster_to_info)

    print("\n[7/7] Exporting to Excel...")
    stem = Path(input_path).stem if not Path(input_path).is_dir() else "merged"
    export_to_excel(df, output_path, cluster_to_info, input_stem=stem)

    return df


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else None

    print(f"\n  Input  : {input_path}")
    if output_path:
        print(f"  Output : {output_path}")

    try:
        run_pipeline(input_path, output_path)
        print("\n✓ Pipeline complete!")
    except Exception as e:
        import traceback
        print(f"\n[ERROR] {type(e).__name__}: {e}")
        traceback.print_exc()
        sys.exit(1)